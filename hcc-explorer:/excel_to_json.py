"""
excel_to_json.py
----------------
Canonical data pipeline for the KGRT HCC Research Explorer.

Reads:  Processedraw.xlsx  (the single source-of-truth spreadsheet)
Writes: public/data/responses.json  (flat array consumed by the frontend)

Usage:
    python excel_to_json.py
    python excel_to_json.py --input path/to/Processedraw.xlsx
    python excel_to_json.py --input Processedraw.xlsx --output public/data/responses.json

Re-run this script any time the Excel file changes. Commit both the
Excel and the generated JSON. Do not hand-edit responses.json.
"""

import json
import sys
import argparse
from pathlib import Path
from collections import Counter, defaultdict

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Run: pip install openpyxl")
    sys.exit(1)


# ─── Constants ────────────────────────────────────────────────────────────────

EXPECTED_ROW_COUNT = 324

EXPECTED_HEADERS = ("Q#", "Topic", "Language", "Therapy", "prompt", "model", "repetition", "response")

ALLOWED_THERAPIES = {"General", "RFA", "MWA", "TACE", "TARE", "Y90"}

ALLOWED_LANGUAGES = {"English", "Spanish"}

ALLOWED_MODELS = {"claude-sonnet-4-6", "gpt-5.2"}

ALLOWED_REPETITIONS = {1, 2, 3}

# Canonical mapping: therapy slot in Q# code → therapy label in Therapy column
SLOT_TO_THERAPY = {
    "TG": "General",
    "T1": "RFA",
    "T2": "MWA",
    "T3": "TACE",
    "T4": "TARE",
    "T5": "Y90",
}

# Language suffix in Q# code → Language column value
LANG_SUFFIX_TO_LABEL = {
    "E": "English",
    "S": "Spanish",
}

# Q number → canonical topic label (derived from data, kept here for validation)
QNUM_TO_TOPIC = {
    "Q1": "Treatment Landscape",
    "Q2": "Doctor Options",
    "Q3": "Therapy Definition",
    "Q4": "Mechanism of Action",
    "Q5": "Curative Potential",
    "Q6": "Benefits and Limitations",
    "Q7": "Evidence and Outcomes",
}

# Human-readable model labels for the frontend
MODEL_LABELS = {
    "claude-sonnet-4-6": "Claude",
    "gpt-5.2":           "GPT",
}

# Therapy group for future use (ablation vs transarterial vs general)
THERAPY_GROUP = {
    "General": "general",
    "RFA":     "ablation",
    "MWA":     "ablation",
    "TACE":    "transarterial",
    "TARE":    "transarterial",
    "Y90":     "transarterial",
}

# Q number → whether it requires a therapy (True = therapy-specific, False = general)
QNUM_IS_THERAPY_SPECIFIC = {
    "Q1": False,
    "Q2": False,
    "Q3": True,
    "Q4": True,
    "Q5": True,
    "Q6": True,
    "Q7": True,
}


# ─── Validation ───────────────────────────────────────────────────────────────

def validate(rows: list[tuple]) -> list[str]:
    """
    Run all validation checks on the raw rows (excluding the header).
    Returns a list of error strings. Empty list = all clear.
    """
    errors = []

    # 1. Row count
    if len(rows) != EXPECTED_ROW_COUNT:
        errors.append(
            f"ROW COUNT: Expected {EXPECTED_ROW_COUNT} rows, found {len(rows)}."
        )

    # 2. Build ID candidates and track duplicates
    id_counter = Counter()
    for row in rows:
        q_code, _, _, _, _, model, rep, _ = row
        model_short = "claude" if "claude" in str(model) else "gpt"
        row_id = f"{q_code}_{model_short}_{rep}"
        id_counter[row_id] += 1

    dupes = {k: v for k, v in id_counter.items() if v > 1}
    if dupes:
        for id_val, count in dupes.items():
            errors.append(f"DUPLICATE ID: '{id_val}' appears {count} times.")

    # 3. Per-row field validation
    for i, row in enumerate(rows, start=2):  # row 2 = first data row in Excel
        if len(row) != 8:
            errors.append(f"Row {i}: Expected 8 columns, found {len(row)}.")
            continue

        q_code, topic, language, therapy, prompt, model, repetition, response = row

        # 3a. No empty values
        for col_name, val in zip(EXPECTED_HEADERS, row):
            if val is None or str(val).strip() == "":
                errors.append(f"Row {i}: Column '{col_name}' is empty.")

        # 3b. Q# code format: must be like Q3_T1_E
        parts = str(q_code).split("_")
        if len(parts) != 3:
            errors.append(f"Row {i}: Q# code '{q_code}' does not match expected format Q#_T#_L.")
            continue  # can't do further Q-code checks

        q_num, t_slot, lang_suffix = parts

        # 3c. Q number is valid
        if q_num not in QNUM_TO_TOPIC:
            errors.append(f"Row {i}: Unknown Q number '{q_num}' in code '{q_code}'.")

        # 3d. Therapy slot is valid
        if t_slot not in SLOT_TO_THERAPY:
            errors.append(f"Row {i}: Unknown therapy slot '{t_slot}' in code '{q_code}'.")

        # 3e. Language suffix is valid
        if lang_suffix not in LANG_SUFFIX_TO_LABEL:
            errors.append(f"Row {i}: Unknown language suffix '{lang_suffix}' in code '{q_code}'.")

        # 3f. Therapy column matches slot in Q# code
        expected_therapy = SLOT_TO_THERAPY.get(t_slot)
        if expected_therapy and therapy != expected_therapy:
            errors.append(
                f"Row {i}: Therapy column '{therapy}' does not match "
                f"slot '{t_slot}' in code '{q_code}' (expected '{expected_therapy}')."
            )

        # 3g. Language column matches suffix in Q# code
        expected_lang = LANG_SUFFIX_TO_LABEL.get(lang_suffix)
        if expected_lang and language != expected_lang:
            errors.append(
                f"Row {i}: Language column '{language}' does not match "
                f"suffix '{lang_suffix}' in code '{q_code}' (expected '{expected_lang}')."
            )

        # 3h. Topic matches Q number
        expected_topic = QNUM_TO_TOPIC.get(q_num)
        if expected_topic and topic != expected_topic:
            errors.append(
                f"Row {i}: Topic '{topic}' does not match expected '{expected_topic}' for {q_num}."
            )

        # 3i. Therapy value is in allowed set
        if therapy not in ALLOWED_THERAPIES:
            errors.append(f"Row {i}: Therapy '{therapy}' is not in allowed set {ALLOWED_THERAPIES}.")

        # 3j. Language value is in allowed set
        if language not in ALLOWED_LANGUAGES:
            errors.append(f"Row {i}: Language '{language}' is not in allowed set {ALLOWED_LANGUAGES}.")

        # 3k. Model value is in allowed set
        if model not in ALLOWED_MODELS:
            errors.append(f"Row {i}: Model '{model}' is not in allowed set {ALLOWED_MODELS}.")

        # 3l. Repetition is in allowed set
        if repetition not in ALLOWED_REPETITIONS:
            errors.append(
                f"Row {i}: Repetition '{repetition}' is not in allowed set {ALLOWED_REPETITIONS}."
            )

        # 3m. General questions must have therapy = General
        if q_num in ("Q1", "Q2") and therapy != "General":
            errors.append(
                f"Row {i}: Q1/Q2 must have therapy='General', found '{therapy}'."
            )

        # 3n. Therapy-specific questions must NOT have therapy = General
        if q_num in ("Q3", "Q4", "Q5", "Q6", "Q7") and therapy == "General":
            errors.append(
                f"Row {i}: {q_num} is therapy-specific but has therapy='General'."
            )

        # 3o. Prompt minimum length (sanity — no truncated prompts)
        if prompt and len(str(prompt).strip()) < 10:
            errors.append(f"Row {i}: Prompt is suspiciously short ({len(str(prompt))} chars).")

        # 3p. Response minimum length (sanity — no empty or stub responses)
        if response and len(str(response).strip()) < 100:
            errors.append(f"Row {i}: Response is suspiciously short ({len(str(response))} chars).")

    # 4. Cross-row: prompt should be consistent for same Q# code across models
    prompt_by_qcode = defaultdict(set)
    for row in rows:
        q_code, _, _, _, prompt, _, _, _ = row
        prompt_by_qcode[q_code].add(str(prompt).strip())

    for q_code, prompts in prompt_by_qcode.items():
        if len(prompts) > 1:
            errors.append(
                f"PROMPT INCONSISTENCY: Q code '{q_code}' has {len(prompts)} "
                f"different prompts across rows (should be identical)."
            )

    # 5. Structural completeness: every expected combination should exist
    #    Expected: 54 unique (q_code, model) combos × 3 reps = 162 per model × 2 models = 324
    combo_counter = Counter()
    for row in rows:
        q_code, _, _, _, _, model, rep, _ = row
        combo_counter[(q_code, model, rep)] += 1

    # Check total combos
    expected_combos = set()
    q_codes_general = [f"Q{n}_TG_{l}" for n in (1, 2) for l in ("E", "S")]
    q_codes_therapy = [
        f"Q{n}_{slot}_{l}"
        for n in (3, 4, 5, 6, 7)
        for slot in ("T1", "T2", "T3", "T4", "T5")
        for l in ("E", "S")
    ]
    all_q_codes = q_codes_general + q_codes_therapy  # 4 + 50 = 54

    for q_code in all_q_codes:
        for model in ALLOWED_MODELS:
            for rep in ALLOWED_REPETITIONS:
                expected_combos.add((q_code, model, rep))

    found_combos = set(combo_counter.keys())
    missing = expected_combos - found_combos
    extra = found_combos - expected_combos

    for combo in sorted(missing):
        errors.append(f"MISSING ROW: ({combo[0]}, {combo[1]}, rep{combo[2]}) not found in data.")

    for combo in sorted(extra):
        errors.append(f"UNEXPECTED ROW: ({combo[0]}, {combo[1]}, rep{combo[2]}) not expected.")

    return errors


# ─── Transform ────────────────────────────────────────────────────────────────

def transform_row(row: tuple) -> dict:
    """
    Convert a single Excel row tuple into a flat JSON-ready record.
    All fields are explicitly named — no positional assumptions downstream.
    """
    q_code, topic, language, therapy, prompt, model, repetition, response = row

    parts = q_code.split("_")
    q_num = parts[0]               # "Q3"
    t_slot = parts[1]              # "T3"
    lang_suffix = parts[2]         # "E"

    model_short = "claude" if "claude" in model else "gpt"

    # Stable, human-readable unique ID for this observation.
    # Format: {q_code}_{model_short}_{rep}
    # Example: Q3_T3_S_gpt_2
    row_id = f"{q_code}_{model_short}_{repetition}"

    return {
        # ── Identity ──────────────────────────────────────────────────────
        "id":           row_id,          # Stable unique key. Never changes once set.
        "q_code":       q_code,          # Full encoded key: Q3_T1_E
        "q_num":        q_num,           # Question number: Q1–Q7
        "t_slot":       t_slot,          # Therapy slot: TG, T1–T5 (useful for future joins)

        # ── Content dimensions ────────────────────────────────────────────
        "topic":        topic,           # Human label: "Therapy Definition"
        "language":     language,        # "English" | "Spanish"
        "therapy":      therapy,         # "RFA" | "MWA" | "TACE" | "TARE" | "Y90" | "General"
        "therapy_group": THERAPY_GROUP.get(therapy, "unknown"),
                                         # "ablation" | "transarterial" | "general"
                                         # Pre-computed grouping for filter UI and future analysis

        # ── Model info ────────────────────────────────────────────────────
        "model":        model,           # Full model string: "claude-sonnet-4-6"
        "model_label":  MODEL_LABELS.get(model, model),
                                         # Short display label: "Claude" | "GPT"
        "repetition":   int(repetition), # 1 | 2 | 3

        # ── Content ───────────────────────────────────────────────────────
        "prompt":       str(prompt).strip(),
        "response":     str(response).strip(),

        # ── Derived scalars for future joins ──────────────────────────────
        "response_length": len(str(response).strip()),
                                         # Character count. Ready to sort/filter by length.
                                         # Avoids recomputing in the browser for 324 rows.
        "lang_code":    lang_suffix,     # "E" | "S" — compact form for URL params / joins
        "is_general":   therapy == "General",
                                         # Boolean shortcut — avoids string comparison in JS
    }


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Export Processedraw.xlsx → responses.json")
    parser.add_argument(
        "--input", "-i",
        default="Processedraw.xlsx",
        help="Path to the source Excel file (default: Processedraw.xlsx)"
    )
    parser.add_argument(
        "--output", "-o",
        default="public/data/responses.json",
        help="Path to write responses.json (default: public/data/responses.json)"
    )
    parser.add_argument(
        "--skip-validation",
        action="store_true",
        help="Skip validation checks (not recommended — for debugging only)"
    )
    args = parser.parse_args()

    input_path  = Path(args.input)
    output_path = Path(args.output)

    # ── Load ──────────────────────────────────────────────────────────────────
    if not input_path.exists():
        print(f"ERROR: Input file not found: {input_path}")
        sys.exit(1)

    print(f"Reading: {input_path}")
    wb = openpyxl.load_workbook(input_path)

    # The canonical sheet is the only one, but be explicit
    sheet_name = "Results - No Word Limit - v4"
    if sheet_name not in wb.sheetnames:
        # Fall back to first sheet with a warning
        sheet_name = wb.sheetnames[0]
        print(f"WARNING: Expected sheet '{sheet_name}' not found. Using first sheet: '{sheet_name}'")

    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))

    # ── Verify headers ────────────────────────────────────────────────────────
    headers = all_rows[0]
    if headers != EXPECTED_HEADERS:
        print(f"ERROR: Unexpected headers.")
        print(f"  Expected: {EXPECTED_HEADERS}")
        print(f"  Found:    {headers}")
        sys.exit(1)

    data_rows = all_rows[1:]
    print(f"Found {len(data_rows)} data rows.")

    # ── Validate ──────────────────────────────────────────────────────────────
    if not args.skip_validation:
        print("Validating...")
        errors = validate(data_rows)
        if errors:
            print(f"\n{'='*60}")
            print(f"VALIDATION FAILED — {len(errors)} error(s) found:")
            print(f"{'='*60}")
            for err in errors:
                print(f"  ✗  {err}")
            print(f"{'='*60}")
            print("Fix the errors above and re-run. No output file was written.")
            sys.exit(1)
        else:
            print(f"  ✓ All {len(data_rows)} rows valid.")
    else:
        print("WARNING: Validation skipped.")

    # ── Transform ─────────────────────────────────────────────────────────────
    print("Transforming rows...")
    records = [transform_row(row) for row in data_rows]

    # Sort deterministically: q_code → model → repetition
    records.sort(key=lambda r: (r["q_code"], r["model"], r["repetition"]))

    # ── Write ─────────────────────────────────────────────────────────────────
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)

    output_size_kb = output_path.stat().st_size / 1024
    print(f"Written: {output_path}  ({output_size_kb:.1f} KB, {len(records)} records)")
    print("Done.")


if __name__ == "__main__":
    main()
